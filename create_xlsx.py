import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

df = pd.read_csv(r"C:\Users\Thiag\Downloads\leads_duplicados_comercial1_vs_comercial2.tsv", sep='\t', dtype=str)
output = r"C:\Users\Thiag\OneDrive\Área de Trabalho\Analise de Mensagens\leads_duplicados_comercial1_vs_comercial2.xlsx"
df.to_excel(output, index=False, sheet_name='Leads Duplicados')

wb = load_workbook(output)
ws = wb.active

header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
header_fill = PatternFill('solid', fgColor='2F5496')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
data_font = Font(name='Arial', size=10)
thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)
alt_fill = PatternFill('solid', fgColor='F2F2F2')

for col_idx in range(1, ws.max_column + 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

for row_idx in range(2, ws.max_row + 1):
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = data_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center')
        if row_idx % 2 == 0:
            cell.fill = alt_fill

col_widths = {'A': 20, 'B': 28, 'C': 28, 'D': 45, 'E': 45, 'F': 22, 'G': 22, 'H': 12, 'I': 12}
for col, width in col_widths.items():
    ws.column_dimensions[col].width = width

ws.auto_filter.ref = ws.dimensions
ws.freeze_panes = 'A2'
ws.sheet_properties.tabColor = '2F5496'

wb.save(output)
print(f"Created {output} with {ws.max_row - 1} records")
